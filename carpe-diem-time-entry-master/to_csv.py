# -*- coding: utf-8 -*-
"""
This code takes in a specifically formatted excel spreadsheet
and converts the entries into a csv to then submit to carpe diem
"""

import csv
import datetime
from dateutil import parser
from tkinter import Tk
from tkinter import filedialog
from openpyxl import load_workbook


# Setting variables
from_date = ""
to_date = ""
dates = []
hours = []
clients = []
matters = []
jurisdictions = []
tasks = []
descriptions = []


def set_date_range():
    global from_date
    global to_date

    print('Please specify the date range to import (From/To).')
    from_date = get_from_date()
    to_date = get_to_date()
    if False == confirm("Please confirm the date range: " + human_date(from_date) + " - " + human_date(to_date)):
        exit()
        
def get_from_date():
    today = datetime.date.today()
    if today.weekday() == 0:
        last_monday = today - datetime.timedelta(days=7)
    else:
        last_monday = today - datetime.timedelta(days=today.weekday())
    example_date_input = human_date(last_monday)
    user_date = input('From date: (i.e. ' + example_date_input + '): ')
    if "" == user_date:
        user_date = example_date_input
    return parser.parse(user_date)

def get_to_date():
    today = datetime.date.today()
    example_date_input = human_date(today)
    user_date = input('To date: (i.e. ' + example_date_input + '): ')
    if "" == user_date:
        user_date = example_date_input
    return parser.parse(user_date)

def human_date(date):
    return date.strftime('%a, %b') + ' ' + date.strftime('%d').lstrip('0')

def confirm(question):
    response = input(question + " [y/n] ")
    return response.lower() in ["y", ""]

def create_csv():
    global dates
    global hours 
    global clients 
    global matters 
    global jurisdictions 
    global tasks 
    global descriptions
    
    with open('time_to_submit.csv', 'w') as csvfile:
        filewriter = csv.writer(csvfile, delimiter=',')
        
        filewriter.writerow(['Time Entry csv', 'Date', 'Hours', 'Client', 'Matter Code', 'Jurisdiction', 'Task Code', 'Description'])
        k = 0
        for k in range(0, len(dates)):
            filewriter.writerow(['', dates[k], hours[k], clients[k], 
                                 matters[k], jurisdictions[k],tasks[k], 
                                 descriptions[k]])
        print("Your entries have been added to time_to_submit.csv. Please run 'python2 import.py' to submit the entries to Carpe Diem")

def get_excel_data():
    global dates
    global hours 
    global clients 
    global matters 
    global jurisdictions 
    global tasks 
    global descriptions 
    global from_date
    global to_date
 
    #Opening the excel file
    print('Opening dialog where you can choose the Excel file to import...')
    Tk().withdraw()
    file_path = filedialog.askopenfilename(defaultextension=".xlsm", filetypes=(("Excel file", "*.xlsm"),("All Files", "*.*") ))
    wb = load_workbook(file_path, read_only=True)
    
    
    #Setting the sheet to the time entry sheet
    if "Time_Entry" in wb.sheetnames:
        ws = wb['Time_Entry'] 
    else:
        print("Please make sure the Excel sheet with your time entries is titled 'Time_Entry' and try again")
        exit()
        
    # Setting variables
    row_count = ws.max_row
    col_V = []
    for j in range(6, row_count):
        col_V.append(ws["V" + str(j)].value)
    cell_num = 6
    i = 0
    count_defaults = 0
    
    # Collecting all information per row
    for m in range(6, row_count):
        entry = ws["V"+str(m)]
        if entry.value:
            entry_date = parser.parse(str(entry.value)[:10])
            # Checking if the date is in range
            if (entry_date >= from_date and entry_date <= to_date):
                    dates.append(str(entry_date)[:10])
                    
                    # Adding hours if cell is populated
                    if ws['W'+str(cell_num)].value:
                        hours.append(float(ws['W'+str(cell_num)].value))
                    else:
                        print("***********************WARNING***********************")
                        print("Please ensure the number of hours is included in every entry. Cell W" + str(cell_num) + " is missing number of hours.")
                        exit()
                        
                    # Adding Client code if populated
                    if ws['X'+str(cell_num)].value:
                        clients.append(str(ws['X'+str(cell_num)].value))
                    else:
                        print("***********************WARNING***********************")
                        print("Please ensure the client code is included in every entry. Cell X" + str(cell_num) + " is missing the client code.")
                        exit()
                        
                    # Adding Matter code if populated
                    if ws['Y'+str(cell_num)].value:
                        matters.append(str(ws['Y'+str(cell_num)].value))
                    else:
                        print("***********************WARNING***********************")
                        print("Please ensure the matter code is included in every entry. Cell Y" + str(cell_num) + " is missing the matter code.")
                        exit()
                        
                    # Adding task code if populated
                    if ws['Z'+str(cell_num)].value:
                        tasks.append(str(ws['Z'+str(cell_num)].value)) 
                    else:
                        tasks.append('')
                    
                    # Adding Jurisdiction - default if blank
                    if ws['AA'+str(cell_num)].value:
                        jurisdictions.append(str(ws['AA'+str(cell_num)].value))                       
                    else:
                        jurisdictions.append(str(ws['G2'].value))
                        count_defaults += 1
                        
                    # Description if populated
                    if ws['AB'+str(cell_num)].value:
                        descriptions.append(str(ws['AB'+str(cell_num)].value))                       
                    else:
                        print("***********************WARNING***********************")
                        print("Please ensure a description is included in every entry. Cell AA" + str(cell_num) + " is missing a description.")
                        exit()
                        
                    # Incrementing the index for valid entries
                    i += 1
                    
        # Incrementing the cell row value
        cell_num += 1     
    
        
    if count_defaults != 0:
        # Display the entries to be added to the csv
        for z in range(0, len(dates)):
            print(str(dates[z]) + ": " + str(hours[z]) + " -- " +  str(matters[z]) 
                  + " -- " + str(clients[z]) + " -- " + str(tasks[z]) + " -- " + 
                  str(jurisdictions[z]) + " -- " + str(descriptions[z]))
        if False == confirm(str(count_defaults) + " entries are using the default jurisdiction "
                            + str(ws['G2'].value) + " and the entries shown will be added to the csv." +"\nWould you like to continue?"):
            exit()
            
    if len(dates) <= 0:
        print("No entries found for" + str(from_date) +" - "+ str(to_date))
        exit()
        
    #
    
    create_csv()   

    
set_date_range()
get_excel_data()
    



